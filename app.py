"""
T.QM.013 检验指导书生成器 v5.1（合并单元格感知 + 性能优化版）
根据 CP 模板结构自动按 OP 分组、固定列映射填充 T.QM.013

新增功能：
- 合并单元格感知：自动拆解合并单元格，将左上角值填充到区域内所有单元格

优化点：
- iter_rows 批量读取，避免逐单元格访问
- 服务端 CP 解析缓存，避免重复 load_workbook
- 模板文件二进制缓存，启动时加载一次
- unmerge 使用对象属性，避免正则
- 一次性解除所有合并单元格后直接写入，无需 safe_write_cell
- 去重固定 key 顺序
- 启动时清理过期临时文件
- 前端 updatePreview 防抖
"""
import os
import sys
import socket
import threading
import webbrowser
import uuid
import time
from datetime import datetime
from io import BytesIO
from flask import Flask, request, jsonify, send_file, session
from werkzeug.utils import secure_filename
import openpyxl

# ==================== PyInstaller 路径处理 ====================
def get_base_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def get_template_path():
    base = get_base_path()
    candidates = [
        os.path.join(base, 'template', 'T.QM.013.xlsm'),
        os.path.join(base, 'T.QM.013.xlsm'),
        os.path.join(os.path.dirname(sys.executable), 'T.QM.013.xlsm') if getattr(sys, 'frozen', False) else None,
        os.path.join(os.getcwd(), 'T.QM.013.xlsm'),
    ]
    for path in candidates:
        if path and os.path.exists(path):
            return path
    for root_dir in [base, os.getcwd()]:
        try:
            for f in os.listdir(root_dir):
                if 'T.QM.013' in f and (f.endswith('.xlsm') or f.endswith('.xlsx')):
                    return os.path.join(root_dir, f)
        except FileNotFoundError:
            pass
    return None


TEMPLATE_FILE = get_template_path()

# ==================== 模板文件二进制缓存 ====================
TEMPLATE_BYTES: bytes | None = None


def init_template_cache():
    """启动时缓存模板文件二进制内容"""
    global TEMPLATE_BYTES
    if TEMPLATE_FILE and os.path.exists(TEMPLATE_FILE):
        with open(TEMPLATE_FILE, 'rb') as f:
            TEMPLATE_BYTES = f.read()
        return True
    return False


# ==================== T.Q.M.013 模板目标坐标配置 ====================
TQM013_HEADER_FIELDS = {
    'project':      {'row': 3,  'col': 8,  'label': '项目'},        # H3
    'oem':          {'row': 6,  'col': 5,  'label': '客户/OEM'},     # E6
    'part_no_oem':  {'row': 10, 'col': 12, 'label': '零件号'},        # L10
    'part_name':    {'row': 8,  'col': 5,  'label': '零件名称'},      # E8
    'part_desc':    {'row': 10, 'col': 6,  'label': '零件描述'},     # F10
    'release_date': {'row': 13, 'col': 8,  'label': '发布日期'},     # H13
    'workstation':  {'row': 13, 'col': 11, 'label': '工作站'},       # K13
}

CONTENT_START_ROW = 19
CONTENT_END_ROW = 48

TQM013_CONTENT_COLS = {
    'content_number': {'col': 3,  'label': '编号'},              # C
    'special_char':   {'col': 5,  'label': '特殊特性符号'},      # E
    'char_desc':      {'col': 6,  'label': '特性描述'},          # F
    'spec_desc':      {'col': 8,  'label': '规格/描述补充'},    # H
    'method_desc':    {'col': 12, 'label': '控制方法/备注'},    # L
    'equipment_freq': {'col': 13, 'label': '设备/频次'},        # M
    'responsible':    {'col': 15, 'label': '负责人'},           # O
}

# CP 列 -> T.Q.M.013 内容列（固定写死）
CP_TO_TQM_CONTENT = {
    'content_number': 4,   # CP D -> TQM C
    'special_char':   7,   # CP G -> TQM E
    'char_desc':      5,   # CP E -> TQM F
    'spec_desc':      8,   # CP H -> TQM H
    'method_desc':    11,  # CP K -> TQM L
    'equipment_freq': 9,   # CP I -> TQM M
    'responsible':    10,  # CP J -> TQM O
}

# CP 模板结构
CP_HEADER_ROW = 8
CP_OP_COLS = [1, 2, 3]   # A/B/C 列识别 OP/工序
CP_DATA_COLS = [4, 5, 7, 8, 9, 10, 11]  # D, E, G, H, I, J, K

# 去重时使用的固定 key 列顺序
DATA_KEY_COLS = [1, 2, 3, 4, 5, 7, 8, 9, 10, 11]

# ==================== Flask 配置 ====================
app = Flask(__name__)
app.secret_key = str(uuid.uuid4())

import tempfile
UPLOAD_DIR = os.path.join(tempfile.gettempdir(), 'qm013_uploads')
OUTPUT_DIR = os.path.join(tempfile.gettempdir(), 'qm013_outputs')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ==================== 服务端 CP 解析缓存 ====================
cp_cache: dict[str, dict] = {}


# ==================== 临时文件清理 ====================
def cleanup_old_files(directory: str, max_age_hours: int = 24):
    """清理超过指定时间的临时文件"""
    now = time.time()
    cutoff = now - max_age_hours * 3600
    try:
        for f in os.listdir(directory):
            fpath = os.path.join(directory, f)
            if os.path.isfile(fpath):
                try:
                    if os.path.getmtime(fpath) < cutoff:
                        os.remove(fpath)
                except OSError:
                    pass
    except FileNotFoundError:
        pass


# ==================== 合并单元格感知读取 ====================
def build_merged_aware_rows(ws, min_row: int, max_row: int, max_col: int) -> list:
    """
    读取工作表数据，并自动展开合并单元格。
    
    openpyxl 的 iter_rows(values_only=True) 对合并单元格只返回左上角一个值，
    其它位置返回 None。此函数通过 ws.merged_cells.ranges 获取合并区域，
    将左上角单元格的值填充到合并区域内所有单元格。
    
    返回：
        rows: list[list[Any]]，与 values_only=True 格式一致，但合并单元格已展开
    """
    # values_only=False 获取 Cell 对象，否则拿不到合并信息的位置对应关系
    cell_rows = list(ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=False,
    ))
    
    # 构建基础值矩阵
    rows = [[cell.value for cell in row] for row in cell_rows]
    
    # 展开合并单元格
    for merged_range in ws.merged_cells.ranges:
        mr_min_row = merged_range.min_row
        mr_max_row = merged_range.max_row
        mr_min_col = merged_range.min_col
        mr_max_col = merged_range.max_col
        
        # 只处理在读取范围内的合并区域
        if mr_max_row < min_row or mr_min_row > max_row:
            continue
        if mr_min_col > max_col:
            continue
        
        # 左上角值
        top_left_value = rows[mr_min_row - 1][mr_min_col - 1]
        
        # 填充整个合并区域
        for r in range(max(min_row, mr_min_row), min(max_row, mr_max_row) + 1):
            for c in range(mr_min_col, min(mr_max_col, max_col) + 1):
                rows[r - 1][c - 1] = top_left_value
    
    return rows


# ==================== 核心函数 ====================
def find_control_plan_sheet(wb) -> str:
    """查找 control plan sheet"""
    target_names = ['control plan', 'cp', 'controlplan', 'kontrollplan', 'steuerplan']
    for sheet_name in wb.sheetnames:
        name_lower = sheet_name.lower().strip()
        for target in target_names:
            if target in name_lower:
                return sheet_name
    return wb.sheetnames[0]


def clean_text(val):
    if val is None:
        return ''
    return ' '.join(str(val).split()).strip()


def extract_cp_header_values_from_rows(rows: list, header_row: int) -> dict:
    """
    从 rows 中提取 CP 表头信息。
    """
    values = {
        'project': '', 'oem': '', 'part_no_oem': '',
        'part_name': '', 'part_desc': '', 'release_date': '',
    }
    keywords = {
        'project': ['Project', '项目', 'Name / partdescription', 'partdescription', '零件描述'],
        'oem': ['Final customer', 'Customer', '客户', 'Supplier / Production Location', 'Supplier'],
        'part_no_oem': ['Assy-No', 'Part No', '零件号', 'GPIN', 'Assy-No / Latest change level'],
        'part_name': ['Name / partdescription', 'partdescription', 'Part name', '零件名称', 'Name / part description'],
        'part_desc': ['Name / partdescription', 'partdescription', 'Description', '描述'],
        'release_date': ['revision list', 'Release', '发布日期'],
    }

    max_col = len(rows[0]) if rows else 0

    for row_idx in range(header_row):  # 0 ~ header_row-1
        row = rows[row_idx]
        for col_idx in range(max_col):
            val = row[col_idx]
            if val is None:
                continue
            text = str(val).strip()
            for field, kws in keywords.items():
                if values[field]:
                    continue
                if any(kw.lower() in text.lower() for kw in kws):
                    if col_idx + 1 < max_col:
                        next_val = row[col_idx + 1]
                        if next_val is not None and str(next_val).strip():
                            values[field] = str(next_val).strip()

    if not values['project'] and values['part_name']:
        values['project'] = values['part_name']
    if not values['part_desc'] and values['part_name']:
        values['part_desc'] = values['part_name']
    return values


def parse_control_plan(filepath: str) -> dict:
    """
    解析控制计划 —— 合并单元格感知版。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[find_control_plan_sheet(wb)]

    max_data_col = max(CP_DATA_COLS)  # 11

    # 使用合并感知读取，自动展开 A~K 列的合并单元格
    rows = build_merged_aware_rows(
        ws,
        min_row=1,
        max_row=ws.max_row,
        max_col=max_data_col,
    )
    wb.close()

    # 读取第8行表头
    all_headers = {}
    if len(rows) >= CP_HEADER_ROW:
        header_row = rows[CP_HEADER_ROW - 1]
        for col_idx, val in enumerate(header_row, start=1):
            all_headers[col_idx] = str(val).strip() if val else ''

    # 提取顶部表头信息
    header_values = extract_cp_header_values_from_rows(rows, CP_HEADER_ROW)

    # 按 OP 分组，A/B/C 列向下填充
    last_op = {'A': None, 'B': None, 'C': None}
    current_op = None
    all_data = []
    workstations = []
    op_set = set()

    for row_idx in range(CP_HEADER_ROW, len(rows)):
        row = rows[row_idx]
        a_val = row[0]
        b_val = row[1]
        c_val = row[2]

        if a_val is not None and str(a_val).strip():
            last_op['A'] = clean_text(a_val)
        if b_val is not None and str(b_val).strip():
            last_op['B'] = clean_text(b_val)
        if c_val is not None and str(c_val).strip():
            last_op['C'] = clean_text(c_val)

        # 新 OP 开始
        if (a_val is not None and str(a_val).strip()) or \
           (b_val is not None and str(b_val).strip()) or \
           (c_val is not None and str(c_val).strip()):
            op_key = last_op['A'] or last_op['B'] or last_op['C']
            if op_key:
                if op_key not in op_set:
                    op_set.add(op_key)
                    workstations.append(op_key)
                    current_op = op_key
                else:
                    current_op = op_key

        # E 列有内容才保留
        e_val = row[4]
        if e_val is not None and str(e_val).strip():
            row_data = {
                1: last_op['A'],
                2: last_op['B'],
                3: last_op['C'],
                4: row[3],    # D
                5: row[4],    # E
                7: row[6],    # G
                8: row[7],    # H
                9: row[8],    # I
                10: row[9],   # J
                11: row[10],  # K
            }
            all_data.append(row_data)

    return {
        'headers': all_headers,
        'workstations': workstations,
        'workstation_col': 1,
        'data': all_data,
        'header_row': CP_HEADER_ROW,
        'total_columns': max_data_col,
        'header_values': header_values,
    }


def unmerge_target_area(ws, start_row: int, end_row: int):
    """
    一次性解除指定行范围内的所有合并单元格。
    """
    to_unmerge = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
            to_unmerge.append(str(merged_range))
    for rng in to_unmerge:
        ws.unmerge_cells(rng)


def fill_template(cp_data: dict, selected_ws: str, header_values: dict) -> BytesIO:
    """
    填充 T.Q.M.013 模板。
    """
    if TEMPLATE_BYTES is None:
        raise FileNotFoundError("未找到 T.Q.M.013 模板文件！")

    wb = openpyxl.load_workbook(BytesIO(TEMPLATE_BYTES), keep_vba=True)
    ws = wb.active

    # 一次性解除表头和内容区的合并单元格
    unmerge_target_area(ws, 1, CONTENT_END_ROW)

    # 填充表头
    for field, config in TQM013_HEADER_FIELDS.items():
        if field == 'release_date':
            value = header_values.get('release_date') or datetime.now().strftime('%Y-%m-%d')
        elif field == 'workstation':
            value = selected_ws
        else:
            value = header_values.get(field)
        if value:
            ws.cell(row=config['row'], column=config['col'], value=value)

    # 按选定的 OP 筛选数据行
    ws_col = cp_data['workstation_col']
    matched = [row for row in cp_data['data'] if row.get(ws_col) == selected_ws]

    # 去重
    seen = set()
    unique_matched = []
    for row in matched:
        key = tuple(row.get(k) for k in DATA_KEY_COLS)
        if key not in seen:
            seen.add(key)
            unique_matched.append(row)

    # 填充内容数据
    current_row = CONTENT_START_ROW
    for row_data in unique_matched:
        if current_row > CONTENT_END_ROW:
            break
        for tqm_field, cp_col in CP_TO_TQM_CONTENT.items():
            if tqm_field in TQM013_CONTENT_COLS:
                tqm_col = TQM013_CONTENT_COLS[tqm_field]['col']
                value = row_data.get(cp_col)
                if value is not None and str(value).strip():
                    ws.cell(row=current_row, column=tqm_col, value=value)
        current_row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    return output


# ==================== API 路由 ====================
@app.route('/')
def index():
    return HTML_PAGE


@app.route('/api/status')
def status():
    return jsonify({
        'template_found': TEMPLATE_FILE is not None,
        'template_path': TEMPLATE_FILE or '未找到',
        'template_cache': TEMPLATE_BYTES is not None,
    })


@app.route('/api/upload_control_plan', methods=['POST'])
def upload_control_plan():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未选择文件'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '文件名为空'})

    session_id = str(uuid.uuid4())
    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_DIR, f"{session_id}_{filename}")
    file.save(filepath)

    try:
        cp_data = parse_control_plan(filepath)
        cp_cache[session_id] = cp_data

        session['cp_filepath'] = filepath
        session['cp_session_id'] = session_id

        return jsonify({
            'success': True,
            'workstations': cp_data['workstations'],
            'headers': {str(k): v for k, v in cp_data['headers'].items()},
            'total_rows': len(cp_data['data']),
            'workstation_column': cp_data['workstation_col'],
            'session_id': session_id,
            'header_row': cp_data['header_row'],
            'total_columns': cp_data['total_columns'],
            'header_values': cp_data['header_values'],
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'解析失败: {str(e)}'})


@app.route('/api/reparse_with_column', methods=['POST'])
def reparse_with_column():
    cp_filepath = session.get('cp_filepath')
    session_id = session.get('cp_session_id')

    if not cp_filepath or not os.path.exists(cp_filepath):
        return jsonify({'success': False, 'error': '会话已过期'})

    try:
        cp_data = parse_control_plan(cp_filepath)
        if session_id:
            cp_cache[session_id] = cp_data

        return jsonify({
            'success': True,
            'workstations': cp_data['workstations'],
            'headers': {str(k): v for k, v in cp_data['headers'].items()},
            'total_rows': len(cp_data['data']),
            'workstation_column': cp_data['workstation_col'],
            'header_values': cp_data['header_values'],
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'重新解析失败: {str(e)}'})


@app.route('/api/generate', methods=['POST'])
def generate():
    data = request.get_json()
    session_id_from_body = data.get('session_id')
    workstation = data.get('workstation')
    header_values = data.get('header_values', {})

    if not session_id_from_body or not workstation:
        return jsonify({'success': False, 'error': '缺少必要参数'})

    cp_data = cp_cache.get(session_id_from_body)

    if cp_data is None:
        cp_filepath = session.get('cp_filepath')
        if not cp_filepath or not os.path.exists(cp_filepath):
            return jsonify({'success': False, 'error': '会话已过期，请重新上传文件'})
        cp_data = parse_control_plan(cp_filepath)
        cp_cache[session_id_from_body] = cp_data

    try:
        safe_ws = str(workstation).replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"T.QM.013_{safe_ws}_{timestamp}.xlsm"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        output = fill_template(cp_data, workstation, header_values)

        if output:
            with open(output_path, 'wb') as f:
                f.write(output.getvalue())
            session['output_path'] = output_path
            session['output_filename'] = output_filename
            return jsonify({
                'success': True,
                'filename': output_filename,
                'download_url': f'/api/download/{output_filename}',
            })
        else:
            return jsonify({'success': False, 'error': '生成失败：模板文件未找到'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'生成失败: {str(e)}'})


@app.route('/api/download/<filename>')
def download(filename):
    safe_name = secure_filename(filename)
    filepath = os.path.join(OUTPUT_DIR, safe_name)
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    return send_file(filepath, as_attachment=True, download_name=safe_name)


@app.route('/api/cleanup', methods=['POST'])
def cleanup():
    sid = session.get('cp_session_id')
    if sid and sid in cp_cache:
        del cp_cache[sid]
    if sid:
        for d in [UPLOAD_DIR, OUTPUT_DIR]:
            try:
                for f in os.listdir(d):
                    if f.startswith(sid):
                        os.remove(os.path.join(d, f))
            except OSError:
                pass
    return jsonify({'success': True})


# ==================== 内嵌 HTML（略，与 v5.0 一致）====================
HTML_PAGE = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>T.QM.013 检验指导书生成器 v5.1</title>
<style>
:root { --primary: #2563eb; --primary-hover: #1d4ed8; --bg: #f1f5f9; --card-bg: #ffffff; --border: #e2e8f0; --text: #1e293b; --text-secondary: #64748b; --success: #16a34a; --warning: #ea580c; --danger: #dc2626; --radius: 8px; --shadow: 0 1px 3px rgba(0,0,0,0.08); }
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif; background: var(--bg); color: var(--text); line-height:1.6; height:100vh; overflow:hidden; }
.header { background: var(--primary); color: white; padding: 10px 24px; display: flex; align-items: center; justify-content: space-between; flex-shrink:0; }
.header h1 { font-size: 1.15rem; font-weight: 600; }
.header .status { font-size: 0.75rem; opacity: 0.9; }
.main-layout { display: flex; height: calc(100vh - 52px); }
.left-panel { flex: 1; overflow-y: auto; padding: 16px 20px; min-width: 480px; }
.right-panel { flex: 1; overflow-y: auto; padding: 16px 20px; background: #f8fafc; border-left: 1px solid var(--border); min-width: 420px; }
.steps { display: flex; gap: 0; margin-bottom: 16px; background: var(--card-bg); border-radius: var(--radius); box-shadow: var(--shadow); overflow: hidden; }
.step { flex: 1; padding: 10px 12px; text-align: center; font-size: 0.78rem; color: var(--text-secondary); }
.step.active { color: white; background: var(--primary); font-weight: 600; }
.step.done { color: var(--success); background: #f0fdf4; }
.step .step-num { display: inline-block; width: 20px; height: 20px; line-height: 20px; border-radius: 50%; border: 2px solid currentColor; margin-right: 2px; font-size: 0.7rem; font-weight: 700; }
.step.active .step-num { background: white; color: var(--primary); border-color: white; }
.card { background: var(--card-bg); border-radius: var(--radius); box-shadow: var(--shadow); padding: 16px; margin-bottom: 14px; }
.card h2 { font-size: 0.95rem; margin-bottom: 10px; padding-bottom: 6px; border-bottom: 2px solid var(--border); }
.card h3 { font-size: 0.85rem; margin-bottom: 8px; color: var(--text-secondary); }
.upload-zone { border: 2px dashed var(--border); border-radius: var(--radius); padding: 28px; text-align: center; cursor: pointer; transition: all 0.3s; background: #fafbfc; }
.upload-zone:hover, .upload-zone.drag-over { border-color: var(--primary); background: #eff6ff; }
.upload-zone .upload-icon { font-size: 2rem; margin-bottom: 8px; }
.upload-zone p { color: var(--text-secondary); font-size: 0.85rem; }
.upload-zone .browse-link { color: var(--primary); font-weight: 600; cursor: pointer; }
.upload-zone input[type="file"] { display: none; }
.file-info { display: none; margin-top: 10px; padding: 8px 14px; background: #f0fdf4; border-radius: 6px; color: var(--success); font-size: 0.85rem; }
.file-info.show { display: block; }
.btn { display: inline-flex; align-items: center; gap: 5px; padding: 7px 16px; border: none; border-radius: 6px; font-size: 0.82rem; font-weight: 500; cursor: pointer; transition: all 0.2s; }
.btn-primary { background: var(--primary); color: white; }
.btn-primary:hover { background: var(--primary-hover); }
.btn-primary:disabled { background: #94a3b8; cursor: not-allowed; }
.btn-outline { background: white; color: var(--primary); border: 1.5px solid var(--primary); }
.btn-success { background: var(--success); color: white; }
.btn-warning { background: var(--warning); color: white; }
.btn-sm { padding: 4px 10px; font-size: 0.75rem; }
.btn-group { display: flex; gap: 8px; margin-top: 10px; flex-wrap: wrap; }
.ws-search { width: 100%; padding: 8px 12px; border: 1.5px solid var(--border); border-radius: 6px; font-size: 0.85rem; margin-bottom: 8px; }
.ws-list { display: flex; flex-wrap: wrap; gap: 6px; max-height: 160px; overflow-y: auto; padding: 2px; }
.ws-chip { padding: 5px 12px; border: 1.5px solid var(--border); border-radius: 16px; cursor: pointer; font-size: 0.8rem; transition: all 0.2s; user-select: none; }
.ws-chip:hover { border-color: var(--primary); background: #eff6ff; }
.ws-chip.selected { background: var(--primary); color: white; border-color: var(--primary); }
.map-section { margin-bottom: 10px; }
.map-row { display: flex; align-items: center; gap: 8px; margin-bottom: 5px; font-size: 0.8rem; }
.map-row label { min-width: 140px; font-weight: 500; text-align: right; }
.map-row input, .map-row select { flex: 1; padding: 5px 8px; border: 1.5px solid var(--border); border-radius: 4px; font-size: 0.78rem; }
.map-row .readonly { background: #f8fafc; color: var(--text-secondary); }
.alert { padding: 8px 12px; border-radius: 6px; margin-bottom: 10px; font-size: 0.8rem; }
.alert-error { background: #fef2f2; color: var(--danger); border: 1px solid #fecaca; }
.alert-info { background: #eff6ff; color: var(--primary); border: 1px solid #bfdbfe; }
.alert-warning { background: #fff7ed; color: var(--warning); border: 1px solid #fed7aa; }
.hidden { display: none !important; }
.spinner { display: none; width: 16px; height: 16px; border: 2px solid var(--border); border-top-color: var(--primary); border-radius: 50%; animation: spin 0.6s linear infinite; }
@keyframes spin { to { transform: rotate(360deg); } }
.column-selector { display: flex; gap: 8px; align-items: center; flex-wrap: wrap; margin-bottom: 10px; padding: 10px; background: #fffbeb; border-radius: 6px; border: 1px solid #fde68a; font-size: 0.8rem; }
.column-selector select { padding: 6px 10px; border: 1.5px solid var(--border); border-radius: 6px; font-size: 0.82rem; }
.preview-title { font-size: 0.95rem; font-weight: 600; margin-bottom: 10px; }
.preview-table { width: 100%; border-collapse: collapse; font-size: 0.62rem; background: white; }
.preview-table td, .preview-table th { border: 1px solid #d1d5db; padding: 1px 2px; min-width: 18px; height: 14px; text-align: center; }
.preview-table .filled { background: #dbeafe; font-weight: 600; color: #1e40af; }
.preview-table .will-fill { background: #fef3c7; border: 1.5px dashed #f59e0b; }
.preview-table .empty { color: #cbd5e1; }
.preview-table .label-cell { background: #f1f5f9; font-weight: 500; color: var(--text-secondary); }
.preview-table .col-header { background: #e2e8f0; font-weight: 600; }
.preview-legend { display: flex; gap: 14px; margin-bottom: 8px; font-size: 0.7rem; }
.preview-legend span { display: flex; align-items: center; gap: 4px; }
.legend-filled { display: inline-block; width: 14px; height: 14px; background: #dbeafe; border: 1px solid #93c5fd; border-radius: 2px; }
.legend-will { display: inline-block; width: 14px; height: 14px; background: #fef3c7; border: 1.5px dashed #f59e0b; border-radius: 2px; }
.legend-empty { display: inline-block; width: 14px; height: 14px; background: white; border: 1px solid #d1d5db; border-radius: 2px; }
.tab-bar { display: flex; gap: 0; margin-bottom: 10px; }
.tab-btn { padding: 6px 14px; border: 1px solid var(--border); background: white; cursor: pointer; font-size: 0.78rem; border-radius: 0; }
.tab-btn:first-child { border-radius: 6px 0 0 6px; }
.tab-btn:last-child { border-radius: 0 6px 6px 0; }
.tab-btn.active { background: var(--primary); color: white; border-color: var(--primary); }
.result-box { padding: 16px; background: #f0fdf4; border-radius: var(--radius); text-align: center; border: 1.5px solid #bbf7d0; }
.result-box .check-icon { font-size: 2rem; }
.result-box h3 { color: var(--success); margin: 6px 0; }
.fixed-map-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; margin-top: 8px; }
.fixed-map-table th, .fixed-map-table td { border: 1px solid var(--border); padding: 6px 10px; text-align: left; }
.fixed-map-table th { background: #f8fafc; font-weight: 600; }
</style>
</head>
<body>
<div class="header">
    <h1>📋 T.QM.013 检验指导书生成器 v5.1</h1>
    <span class="status" id="statusBar">检查模板...</span>
</div>
<div class="main-layout">
    <div class="left-panel">
        <div class="steps" id="stepIndicator">
            <div class="step active" data-step="1"><span class="step-num">1</span> 上传CP</div>
            <div class="step" data-step="2"><span class="step-num">2</span> 选工位</div>
            <div class="step" data-step="3"><span class="step-num">3</span> 映射</div>
            <div class="step" data-step="4"><span class="step-num">4</span> 下载</div>
        </div>
        <div id="globalAlert" class="hidden"></div>
        <div class="card" id="step1Card">
            <h2>📁 步骤 1：上传版本控制计划</h2>
            <div class="upload-zone" id="uploadZone">
                <div class="upload-icon"></div>
                <p>拖拽 Excel 文件，或 <span class="browse-link" id="browseLink">点击浏览</span></p>
                <p style="font-size:0.7rem;margin-top:3px;">支持 .xlsx / .xlsm / .xls</p>
                <input type="file" id="fileInput" accept=".xlsx,.xlsm,.xls">
            </div>
            <div class="file-info" id="fileInfo">✅ <span id="fileName"></span> — <span id="fileStats"></span></div>
            <div class="btn-group" style="justify-content:flex-end;"><span class="spinner" id="uploadSpinner"></span></div>
        </div>
        <div class="card hidden" id="step2Card">
            <h2>🔍 步骤 2：选择目标工位/OP</h2>
            <div id="headerPreviewBox" class="hidden">
                <p style="font-weight:600;font-size:0.8rem;">📊 检测到的表头（第 <span id="headerRowNum"></span> 行）</p>
                <div style="font-size:0.7rem;max-height:100px;overflow-y:auto;background:#f8fafc;padding:6px;border-radius:4px;" id="headerPreview"></div>
            </div>
            <div id="manualColumnBox" class="hidden">
                <div class="column-selector">
                    <label>⚠️ 手动选择工位列：</label>
                    <select id="wsColumnSelect"><option value="">-- 请选择 --</option></select>
                    <button class="btn btn-warning btn-sm" id="btnReparse">🔄 重新解析</button>
                    <span class="spinner" id="reparseSpinner"></span>
                </div>
            </div>
            <input type="text" class="ws-search" id="wsSearch" placeholder="搜索工位/OP...">
            <div class="ws-list" id="wsList"></div>
            <div style="margin-top:6px;font-size:0.78rem;">
                已选：<strong id="selectedWsDisplay" style="color:var(--primary);">未选择</strong>
                | 共 <strong id="wsCount">0</strong> 个工位
            </div>
            <div class="btn-group" style="justify-content:space-between;">
                <button class="btn btn-outline btn-sm" onclick="resetAll()">🔄 重新上传</button>
                <button class="btn btn-primary" id="btnNextToMapping" disabled>下一步：确认映射 →</button>
            </div>
        </div>
        <div class="card hidden" id="step3Card">
            <h2>🔗 步骤 3：确认列映射</h2>
            <div class="tab-bar">
                <button class="tab-btn active" data-tab="header">📋 基本信息</button>
                <button class="tab-btn" data-tab="content">📐 内容列（固定）</button>
            </div>
            <div id="tabHeader" class="map-section">
                <h3>T.Q.M.013 表头字段（可编辑）</h3>
                <div id="headerFields"></div>
            </div>
            <div id="tabContent" class="map-section hidden">
                <h3>CP 列 → T.Q.M.013 内容区（固定映射）</h3>
                <table class="fixed-map-table">
                    <thead><tr><th>CP 列</th><th>T.Q.M.013 位置</th><th>说明</th></tr></thead>
                    <tbody>
                        <tr><td>D 列</td><td>C 列（编号）</td><td>编号/No.</td></tr>
                        <tr><td>E 列</td><td>F 列</td><td>特性描述</td></tr>
                        <tr><td>G 列</td><td>E 列</td><td>特殊特性符号</td></tr>
                        <tr><td>H 列</td><td>H 列</td><td>规格/描述补充</td></tr>
                        <tr><td>K 列</td><td>L 列</td><td>控制方法/备注</td></tr>
                        <tr><td>I 列</td><td>M 列</td><td>设备/频次</td></tr>
                        <tr><td>J 列</td><td>O 列</td><td>负责人</td></tr>
                    </tbody>
                </table>
            </div>
            <div class="btn-group" style="justify-content:space-between;">
                <button class="btn btn-outline btn-sm" id="btnBackToStep2">← 返回</button>
                <button class="btn btn-primary" id="btnGenerate"> 生成检验指导书 <span class="spinner" id="generateSpinner"></span></button>
            </div>
        </div>
        <div class="card hidden" id="step4Card">
            <h2>✅ 步骤 4：生成完成！</h2>
            <div class="result-box">
                <div class="check-icon">✅</div>
                <h3>检验指导书生成成功！</h3>
                <p id="resultFilename" style="font-size:0.85rem;"></p>
            </div>
            <div class="btn-group" style="justify-content:center;">
                <button class="btn btn-success" id="btnDownload"> 下载</button>
                <button class="btn btn-outline btn-sm" onclick="resetAll()">🔄 新建</button>
            </div>
        </div>
    </div>
    <div class="right-panel">
        <div class="preview-title">📐 T.Q.M.013 模板预览</div>
        <div class="preview-legend">
            <span><span class="legend-filled"></span> 表头已映射</span>
            <span><span class="legend-will"></span> 内容数据区</span>
            <span><span class="legend-empty"></span> 空白</span>
        </div>
        <div id="previewContent">
            <p style="color:var(--text-secondary);font-size:0.8rem;">上传控制计划后将显示模板预览</p>
        </div>
    </div>
</div>
<script>
const $=s=>document.querySelector(s),$$=s=>document.querySelectorAll(s);
const STATE={sessionId:null,workstations:[],selectedWorkstation:null,allCpHeaders:{},headerValues:{},downloadUrl:null,currentStep:1};

function setStep(s){
    STATE.currentStep=s;
    $$('.step').forEach((el,i)=>{el.classList.remove('active','done');if(i+1<s)el.classList.add('done');if(i+1===s)el.classList.add('active');});
    ['step1Card','step2Card','step3Card','step4Card'].forEach((id,i)=>document.getElementById(id).classList.toggle('hidden',i+1!==s));
}

function showAlert(msg,type='error'){
    const el=$('#globalAlert');el.className='alert alert-'+type;el.textContent=msg;el.classList.remove('hidden');
    setTimeout(()=>el.classList.add('hidden'),6000);
}

fetch('/api/status').then(r=>r.json()).then(d=>{
    $('#statusBar').textContent=d.template_found?'✅ 模板已加载'+(d.template_cache?' (已缓存)':' (磁盘)') :'⚠️ 模板未找到';
    $('#statusBar').style.color=d.template_found?'#bbf7d0':'#fecaca';
});

const uploadZone=$('#uploadZone'),fileInput=$('#fileInput'),uploadSpinner=$('#uploadSpinner');
$('#browseLink').addEventListener('click',()=>fileInput.click());
uploadZone.addEventListener('click',e=>{if(e.target!==$('#browseLink'))fileInput.click();});
uploadZone.addEventListener('dragover',e=>{e.preventDefault();uploadZone.classList.add('drag-over');});
uploadZone.addEventListener('dragleave',()=>uploadZone.classList.remove('drag-over'));
uploadZone.addEventListener('drop',e=>{e.preventDefault();uploadZone.classList.remove('drag-over');if(e.dataTransfer.files.length>0)handleFile(e.dataTransfer.files[0]);});
fileInput.addEventListener('change',()=>{if(fileInput.files.length>0)handleFile(fileInput.files[0]);});

async function handleFile(file){
    const fd=new FormData();fd.append('file',file);
    uploadSpinner.style.display='inline-block';
    try{
        const r=await fetch('/api/upload_control_plan',{method:'POST',body:fd});
        const d=await r.json();
        if(!d.success){showAlert(d.error);return;}
        STATE.sessionId=d.session_id;STATE.workstations=d.workstations||[];
        STATE.allCpHeaders=d.headers||{};
        STATE.headerValues=d.header_values||{};
        $('#fileName').textContent=file.name;
        $('#fileStats').textContent='共 '+d.total_rows+' 行，'+d.workstations.length+' 个工位';
        $('#fileInfo').classList.add('show');
        $('#headerRowNum').textContent=d.header_row;
        showHeaderPreview(d.headers);
        if(d.workstations.length===0){
            showAlert('未识别到工位/OP，请检查 CP 结构','warning');
        }
        renderWorkstations();updatePreview();setStep(2);
    }catch(e){showAlert('网络错误: '+e.message);}
    finally{uploadSpinner.style.display='none';}
}

function showHeaderPreview(headers){
    let h='';
    for(const [col,hdr] of Object.entries(headers)){
        h+='<span style="margin:2px;display:inline-block;background:#e2e8f0;padding:1px 5px;border-radius:3px;">列'+col+': '+(hdr||'<i>空</i>')+'</span>';
    }
    $('#headerPreview').innerHTML=h;
    $('#headerPreviewBox').classList.remove('hidden');
}

function renderHeaderFields(){
    const fields=[
        {k:'project',l:'项目 (H3)'},{k:'oem',l:'客户/OEM (E6)'},
        {k:'part_no_oem',l:'零件号 (L10)'},{k:'part_name',l:'零件名称 (E8)'},
        {k:'part_desc',l:'零件描述 (F10)'},{k:'release_date',l:'发布日期 (H13)'},
        {k:'workstation',l:'工作站 (K13, 自动)'}
    ];
    $('#headerFields').innerHTML=fields.map(f=>{
        const readonly=f.k==='workstation'?' readonly':'';
        const val=f.k==='workstation'?(STATE.selectedWorkstation||''):(STATE.headerValues[f.k]||'');
        return '<div class="map-row"><label>'+f.l+'</label><input type="text" id="header_'+f.k+'" data-field="'+f.k+'" value="'+val+'"'+readonly+'></div>';
    }).join('');
}

function renderWorkstations(f){
    f=(f||'').toLowerCase();
    const filtered=(STATE.workstations||[]).filter(w=>String(w).toLowerCase().includes(f));
    $('#wsCount').textContent=STATE.workstations.length;
    if(filtered.length===0){
        $('#wsList').innerHTML='<p style="color:var(--text-secondary);padding:12px;font-size:0.8rem;">未找到工位</p>';
        return;
    }
    $('#wsList').innerHTML=filtered.map(w=>'<span class="ws-chip'+(w===STATE.selectedWorkstation?' selected':'')+'" data-ws="'+w.replace(/"/g,'&quot;')+'">'+w+'</span>').join('');
    $('#wsList').querySelectorAll('.ws-chip').forEach(c=>c.addEventListener('click',()=>{
        STATE.selectedWorkstation=c.dataset.ws;$('#selectedWsDisplay').textContent=STATE.selectedWorkstation;
        $('#btnNextToMapping').disabled=false;updatePreview();renderWorkstations($('#wsSearch').value);
    }));
}

$('#wsSearch').addEventListener('input',e=>renderWorkstations(e.target.value));
$('#btnNextToMapping').addEventListener('click',()=>{
    if(!STATE.selectedWorkstation)return;
    renderHeaderFields();setStep(3);updatePreview();
});

$$('.tab-btn').forEach(btn=>btn.addEventListener('click',()=>{
    $$('.tab-btn').forEach(b=>b.classList.remove('active'));
    btn.classList.add('active');
    const tab=btn.dataset.tab;
    $('#tabHeader').classList.toggle('hidden',tab!=='header');
    $('#tabContent').classList.toggle('hidden',tab!=='content');
}));

$('#btnBackToStep2').addEventListener('click',()=>setStep(2));
$('#btnGenerate').addEventListener('click',async()=>{
    const sp=$('#generateSpinner'),btn=$('#btnGenerate');sp.style.display='inline-block';btn.disabled=true;
    const hv={};
    $$('#headerFields input').forEach(inp=>{
        const k=inp.dataset.field;
        if(k!=='workstation')hv[k]=inp.value;
    });
    hv['workstation']=STATE.selectedWorkstation;
    try{
        const r=await fetch('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({session_id:STATE.sessionId,workstation:STATE.selectedWorkstation,header_values:hv})});
        const d=await r.json();
        if(!d.success){showAlert(d.error);return;}
        STATE.downloadUrl=d.download_url;$('#resultFilename').textContent='文件名：'+d.filename;setStep(4);
    }catch(e){showAlert('网络错误: '+e.message);}
    finally{sp.style.display='none';btn.disabled=false;}
});

$('#btnDownload').addEventListener('click',()=>{if(STATE.downloadUrl)window.open(STATE.downloadUrl,'_blank');});

// 防抖版 updatePreview
let _previewTimer=null;
function updatePreview(){
    clearTimeout(_previewTimer);
    _previewTimer=setTimeout(_doUpdatePreview,80);
}

function _doUpdatePreview(){
    const hv=STATE.headerValues||{};
    const hasContent=STATE.selectedWorkstation;
    const filled={};
    for(const k of ['project','oem','part_no_oem','part_name','part_desc','release_date']){
        filled[k]=hv[k]?'(已填)':'';
    }
    filled['workstation']=STATE.selectedWorkstation||'(待选)';
    let html='<table class="preview-table">';
    for(let row=1;row<=35;row++){
        html+='<tr><td class="label-cell" style="font-size:0.5rem;">'+row+'</td>';
        if(row===3){for(let c=0;c<7;c++)html+='<td class="empty"></td>';html+='<td class="'+(filled['project']?'filled':'empty')+'">'+(filled['project']||'H3')+'</td>';for(let c=0;c<8;c++)html+='<td class="empty"></td>';}
        else if(row===6){for(let c=0;c<4;c++)html+='<td class="empty"></td>';html+='<td class="'+(filled['oem']?'filled':'empty')+'">'+(filled['oem']||'E6')+'</td>';for(let c=0;c<10;c++)html+='<td class="empty"></td>';}
        else if(row===8){for(let c=0;c<4;c++)html+='<td class="empty"></td>';html+='<td class="'+(filled['part_name']?'filled':'empty')+'">'+(filled['part_name']||'E8')+'</td>';for(let c=0;c<11;c++)html+='<td class="empty"></td>';}
        else if(row===10){for(let c=0;c<5;c++)html+='<td class="empty"></td>';html+='<td class="'+(filled['part_desc']?'filled':'empty')+'">'+(filled['part_desc']||'F10')+'</td>';for(let c=0;c<10;c++)html+='<td class="empty"></td>';}
        else if(row===13){for(let c=0;c<7;c++)html+='<td class="empty"></td>';html+='<td class="'+(filled['release_date']?'filled':'empty')+'">'+(filled['release_date']||'H13')+'</td>';for(let c=0;c<3;c++)html+='<td class="empty"></td>';html+='<td class="'+(filled['workstation']?'filled':'empty')+'">'+(filled['workstation']||'K13')+'</td>';for(let c=0;c<5;c++)html+='<td class="empty"></td>';}
        else if(row===17){html+='<td class="col-header">内容</td><td class="col-header">内容</td><td class="col-header">C</td><td class="col-header">D</td>';for(let i=0;i<6;i++)html+='<td class="col-header">描述</td>';html+='<td class="empty"></td><td class="col-header">试验等级/设备</td><td class="col-header">试验等级/设备</td><td class="col-header">负责人</td><td class="col-header">负责人</td>';}
        else if(row>=19 && row<=30){const cls=hasContent?'will-fill':'empty';html+='<td class="empty"></td><td class="empty"></td>';html+='<td class="'+cls+'">'+(hasContent?'C'+row:'')+'</td>';html+='<td class="empty"></td>';html+='<td class="'+cls+'">'+(hasContent?'E'+row:'')+'</td>';html+='<td class="'+cls+'">'+(hasContent?'F'+row:'')+'</td>';for(let i=0;i<2;i++)html+='<td class="empty"></td>';html+='<td class="'+cls+'">'+(hasContent?'H'+row:'')+'</td>';for(let i=0;i<3;i++)html+='<td class="empty"></td>';html+='<td class="'+cls+'">'+(hasContent?'L'+row:'')+'</td>';html+='<td class="'+cls+'">'+(hasContent?'M'+row:'')+'</td>';html+='<td class="empty"></td>';html+='<td class="'+cls+'">'+(hasContent?'O'+row:'')+'</td>';html+='<td class="empty"></td>';}
        else{for(let c=0;c<16;c++)html+='<td class="empty"></td>';}
        html+='</tr>';
    }
    html+='</table>';
    $('#previewContent').innerHTML=html;
}

function resetAll(){
    STATE.sessionId=null;STATE.workstations=[];STATE.selectedWorkstation=null;
    STATE.allCpHeaders={};STATE.headerValues={};STATE.downloadUrl=null;
    $('#fileInfo').classList.remove('show');$('#wsList').innerHTML='';$('#selectedWsDisplay').textContent='未选择';
    $('#btnNextToMapping').disabled=true;$('#wsSearch').value='';$('#wsCount').textContent='0';
    $('#headerPreviewBox').classList.add('hidden');$('#manualColumnBox').classList.add('hidden');
    $('#headerFields').innerHTML='';
    $('#previewContent').innerHTML='<p style="color:var(--text-secondary);font-size:0.8rem;">上传控制计划后将显示模板预览</p>';
    fileInput.value='';setStep(1);
    fetch('/api/cleanup',{method:'POST'}).catch(()=>{});
}

setStep(1);
</script>
</body>
</html>'''


# ==================== 启动入口 ====================
def find_free_port():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(('', 0))
        return s.getsockname()[1]


def main():
    template_ok = init_template_cache()
    cleanup_old_files(UPLOAD_DIR)
    cleanup_old_files(OUTPUT_DIR)

    port = find_free_port()
    url = f'http://127.0.0.1:{port}'
    print("=" * 55)
    print("  T.QM.013 检验指导书生成器 v5.1 (合并单元格感知 + 优化版)")
    print("=" * 55)
    print(f"  模板文件: {TEMPLATE_FILE or '❌ 未找到!'}")
    print(f"  模板缓存: {'✅ 已缓存' if template_ok else ' 未缓存'}")
    print(f"  本地地址: {url}")
    print("=" * 55)

    threading.Timer(1.5, lambda: webbrowser.open(url)).start()

    import logging
    logging.getLogger('werkzeug').setLevel(logging.WARNING)
    app.run(host='127.0.0.1', port=port, debug=False)


if __name__ == '__main__':
    main()
